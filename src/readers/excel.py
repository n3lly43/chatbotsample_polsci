"""Reader for Excel files: .xlsx and .xls"""
MAX_CHUNK_CHARS = 6000

def read_excel(file_path: str) -> list[dict]:
    try:
        from pathlib import Path
        ext = Path(file_path).suffix.lower()
        if ext == ".xls":
            return _read_xls(file_path)
        return _read_xlsx(file_path)
    except Exception as e:
        print(f"Warning: Could not read {file_path}: {e}")
        return []

def _read_xlsx(file_path: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        pages = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            pages.extend(_rows_to_pages(rows, sheet_name))
        return pages
    finally:
        wb.close()

def _read_xls(file_path: str) -> list[dict]:
    import xlrd
    wb = xlrd.open_workbook(file_path)
    pages = []
    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        if ws.nrows == 0:
            continue
        rows = []
        for r in range(ws.nrows):
            rows.append(tuple(ws.cell_value(r, c) for c in range(ws.ncols)))
        pages.extend(_rows_to_pages(rows, ws.name))
    return pages

def _rows_to_pages(rows: list[tuple], sheet_name: str) -> list[dict]:
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    row_texts = []
    for row in rows[1:]:
        parts = []
        padded_row = list(row) + [None] * max(0, len(headers) - len(row))
        for header, val in zip(headers, padded_row[:len(headers)]):
            if val is not None and str(val).strip() and str(val).lower() not in ("nan", "nat", "<na>", "inf", "-inf"):
                parts.append(f"{header}: {val}")
        if parts:
            row_texts.append("; ".join(parts))

    header_line = f"Sheet: {sheet_name} | Columns: {', '.join(headers)}\n"
    pages = []
    block = []
    block_chars = len(header_line)
    block_start = 1

    for idx, row_text in enumerate(row_texts):
        if block and block_chars + len(row_text) + 1 > MAX_CHUNK_CHARS:
            text = header_line + "\n".join(block)
            pages.append({"page": f"{sheet_name}_rows_{block_start}-{block_start + len(block) - 1}", "text": text})
            block = []
            block_chars = len(header_line)
            block_start = idx + 1
        block.append(row_text)
        block_chars += len(row_text) + 1

    if block:
        text = header_line + "\n".join(block)
        pages.append({"page": f"{sheet_name}_rows_{block_start}-{block_start + len(block) - 1}", "text": text})
    return pages
