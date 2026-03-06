"""SQL ingestion: load tabular files into SQLite for structured queries."""

import csv
import json
import os
import re
import sqlite3
from pathlib import Path

# Extensions that trigger SQL ingestion (tabular formats)
SQL_EXTENSIONS = {".csv", ".tab", ".tsv", ".xlsx", ".xls", ".dta", ".sav", ".rds", ".rda"}

# Extensions that might contain codebook/documentation
_CODEBOOK_EXTENSIONS = {".pdf", ".docx", ".txt", ".md"}

# Keywords suggesting a file is a codebook
_CODEBOOK_KEYWORDS = {
    "codebook", "readme", "dictionary", "documentation",
    "manual", "guide", "metadata", "variable", "description",
}


def _sanitize_part(s: str) -> str:
    """Sanitize a single name part: replace non-alnum with _, collapse, strip."""
    s = re.sub(r"[^a-zA-Z0-9]", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _sanitize_table_name(dataset: str, stem: str, ext: str, sheet: str = None) -> str:
    """Convert dataset/filename/ext into a valid SQL table identifier."""
    suffix = f"{_sanitize_part(stem)}_{ext.lstrip('.')}"
    if sheet:
        suffix += f"_{_sanitize_part(sheet)}"
    name = f"{_sanitize_part(dataset)}__{suffix}".lower()
    if not name or not name[0].isalpha():
        name = "t_" + name
    return name


def _infer_column_type(values: list) -> str:
    """Infer SQLite column type from a list of raw string values."""
    non_empty = [v for v in values if v is not None and str(v).strip() and str(v).lower() != "nan"]
    if not non_empty:
        return "TEXT"
    # Try INTEGER
    try:
        for v in non_empty:
            int(str(v).strip())
        return "INTEGER"
    except (ValueError, TypeError):
        pass
    # Try REAL
    try:
        for v in non_empty:
            f = float(str(v).strip())
            # Reject inf/nan — these are not valid REAL data
            if not (-1e308 < f < 1e308):
                raise ValueError("non-finite float")
        return "REAL"
    except (ValueError, TypeError):
        pass
    return "TEXT"


def _get_sample_values(values: list, n: int = 5) -> list:
    """Get a representative spread of up to n unique non-empty sample values.

    Instead of taking the first n values (which may be biased by sort order),
    picks evenly spaced values from the sorted unique set so the LLM sees
    the full range (e.g., first, middle, last country names).
    """
    unique = []
    seen = set()
    for v in values:
        if v is not None and str(v).strip() and str(v).lower() != "nan":
            v_str = str(v).strip()
            if v_str not in seen:
                seen.add(v_str)
                unique.append(v_str)
    if not unique:
        return []
    # Sort numerically if all values are numeric, otherwise alphabetically
    try:
        sorted_vals = sorted(unique, key=float)
    except (ValueError, TypeError):
        sorted_vals = sorted(unique)
    if len(sorted_vals) <= n:
        return sorted_vals
    if n <= 1:
        return [sorted_vals[0]]
    # Pick evenly spaced indices including first and last
    indices = [round(i * (len(sorted_vals) - 1) / (n - 1)) for i in range(n)]
    return [sorted_vals[i] for i in indices]


def _get_column_stats(values: list, col_type: str) -> dict:
    """Compute column statistics for schema metadata.

    Returns dict with 'unique_count' and optionally 'min'/'max' for numerics.
    """
    non_empty = [
        str(v).strip() for v in values
        if v is not None and str(v).strip() and str(v).lower() != "nan"
    ]
    unique_count = len(set(non_empty))
    stats = {"unique_count": unique_count}
    if col_type == "INTEGER" and non_empty:
        try:
            nums = [int(v) for v in non_empty]
            stats["min"] = min(nums)
            stats["max"] = max(nums)
        except (ValueError, TypeError):
            pass
    elif col_type == "REAL" and non_empty:
        try:
            nums = [float(v) for v in non_empty]
            stats["min"] = min(nums)
            stats["max"] = max(nums)
        except (ValueError, TypeError):
            pass
    return stats


def _find_codebook_files(tabular_file: Path) -> list[Path]:
    """Search for codebook/documentation files in the same directory as the tabular file."""
    parent = tabular_file.parent
    candidates = []
    for f in sorted(parent.iterdir()):
        if not f.is_file():
            continue
        if f.suffix.lower() not in _CODEBOOK_EXTENSIONS:
            continue
        candidates.append(f)

    # Sort: files with codebook keywords in name come first
    def priority(p):
        name_lower = p.stem.lower()
        return 0 if any(kw in name_lower for kw in _CODEBOOK_KEYWORDS) else 1

    candidates.sort(key=priority)
    return candidates


def _read_codebook_text(codebook_files: list[Path], max_chars: int = 4000) -> str:
    """Read codebook files and return concatenated text, truncated."""
    from src.readers import read_file

    texts = []
    total = 0
    for f in codebook_files:
        try:
            pages = read_file(str(f))
            text = "\n".join(p.get("text", "") for p in pages).strip()
            if text:
                texts.append(f"[From: {f.name}]\n{text}")
                total += len(text)
                if total >= max_chars:
                    break
        except Exception:
            continue

    combined = "\n\n".join(texts)
    if len(combined) > max_chars:
        combined = combined[:max_chars] + "..."
    return combined


def _describe_columns_with_llm(
    table_name: str,
    source_file: str,
    columns_info: list[dict],
    codebook_text: str,
    cfg: dict,
) -> dict:
    """Use LLM to generate column and table descriptions.

    Args:
        table_name: SQL table name.
        source_file: Original file path relative to knowledge base.
        columns_info: Column metadata dicts (name, type, sample, stats).
        codebook_text: Text from codebook files (empty if none found).
        cfg: App config for LLM access.

    Returns:
        Dict with 'table_description' and 'columns' mapping name -> description.
        Empty dict on failure.
    """
    from src.llm import generate

    col_lines = []
    for c in columns_info:
        name = c["name"]
        orig = c.get("original_name", name)
        ctype = c.get("type", "TEXT")
        samples = c.get("sample", [])
        stats = c.get("stats", {})

        parts = [f"- {name}"]
        if orig != name:
            parts.append(f"(originally: {orig})")
        parts.append(f"[{ctype}]")
        if stats.get("unique_count"):
            parts.append(f"({stats['unique_count']} unique)")
        if stats.get("min") is not None:
            parts.append(f"range: {stats['min']}\u2013{stats['max']}")
        if samples:
            parts.append(f"e.g. {', '.join(str(s) for s in samples[:5])}")

        col_lines.append(" ".join(parts))

    cols_detail = "\n".join(col_lines)

    if codebook_text:
        codebook_section = (
            "\nDocumentation/codebook found for this dataset:\n"
            "--- CODEBOOK ---\n"
            f"{codebook_text}\n"
            "--- END CODEBOOK ---\n"
            "Use the codebook to provide accurate column descriptions."
        )
    else:
        codebook_section = (
            "\nNo codebook found. Infer column meanings from names, "
            "types, and sample values. Note what is inferred."
        )

    prompt = (
        f"Analyze this dataset and describe what each column represents.\n\n"
        f"Table: {table_name}\n"
        f"Source: {source_file}\n\n"
        f"Columns:\n{cols_detail}\n"
        f"{codebook_section}\n\n"
        "Return ONLY valid JSON (no markdown fences, no explanation):\n"
        "{\n"
        '  "table_description": "What this table/dataset contains (1-2 sentences)",\n'
        '  "columns": {\n'
        '    "column_name": "What this column represents (1 sentence)"\n'
        "  }\n"
        "}"
    )

    try:
        response = generate(
            "You are a data analyst. Describe dataset columns concisely and accurately.",
            prompt,
            cfg,
            max_tokens=1024,
        )
        text = response.strip()
        start = text.find("{")
        end = text.rfind("}") + 1
        if start >= 0 and end > start:
            return json.loads(text[start:end])
        return {}
    except Exception:
        return {}


def _enrich_schema_with_descriptions(
    schema_registry: dict,
    tabular_files: list[tuple],
    documents_dir: str,
    cfg: dict,
) -> None:
    """Add LLM-generated column descriptions to the schema registry.

    For each table:
    1. Search for codebook/documentation files near the source file.
    2. If found, use LLM + codebook to describe columns.
    3. If not, use LLM + column names/samples to infer descriptions.

    Modifies schema_registry in place.
    """
    # Build mapping from source_file (relative) -> absolute Path
    source_to_path: dict[str, Path] = {}
    for file_path, _dataset_name in tabular_files:
        if file_path.suffix.lower() in SQL_EXTENSIONS:
            try:
                rel_path = str(file_path.relative_to(documents_dir))
                source_to_path[rel_path] = file_path
            except ValueError:
                continue

    for table_name, info in schema_registry.items():
        source_file = info.get("source_file", "")
        file_path = source_to_path.get(source_file)

        # Find and read codebook files
        codebook_text = ""
        if file_path:
            codebook_files = _find_codebook_files(file_path)
            if codebook_files:
                codebook_text = _read_codebook_text(codebook_files)
                print(f"  Found codebook for {table_name}: "
                      f"{[f.name for f in codebook_files]}")

        # Get LLM descriptions
        descriptions = _describe_columns_with_llm(
            table_name, source_file, info["columns"], codebook_text, cfg,
        )

        if descriptions:
            info["table_description"] = descriptions.get("table_description", "")
            col_descs = descriptions.get("columns", {})
            for col in info["columns"]:
                col["description"] = col_descs.get(col["name"], "")
            method = "codebook" if codebook_text else "inferred"
            print(f"  Described {table_name} columns ({method})")
        else:
            print(f"  Could not generate descriptions for {table_name}")


def _load_rows_from_csv(file_path: str, ext: str) -> list[tuple]:
    """Load CSV/TSV/TAB into (sheet_or_name, headers, rows) tuples."""
    delimiter = "\t" if ext in (".tab", ".tsv") else ","
    with open(file_path, "r", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.reader(f, delimiter=delimiter)
        all_rows = list(reader)
    if len(all_rows) < 2:
        return []
    headers = [h.strip().strip('"') for h in all_rows[0]]
    return [(None, headers, all_rows[1:])]


def _load_rows_from_excel(file_path: str, ext: str) -> list[tuple]:
    """Load Excel into (sheet_name, headers, rows) tuples."""
    if ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        results = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            if len(all_rows) < 2:
                continue
            headers = [str(h) if h is not None else "" for h in all_rows[0]]
            rows = [[str(c) if c is not None else "" for c in row] for row in all_rows[1:]]
            results.append((sheet_name, headers, rows))
        wb.close()
        return results
    else:  # .xls
        import xlrd
        wb = xlrd.open_workbook(file_path)
        results = []
        for sheet_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)
            if ws.nrows < 2:
                continue
            headers = [str(ws.cell_value(0, c)) for c in range(ws.ncols)]
            rows = [[str(ws.cell_value(r, c)) for c in range(ws.ncols)] for r in range(1, ws.nrows)]
            results.append((ws.name, headers, rows))
        return results


def _safe_str(v):
    """Convert a value to string, preserving None for missing data."""
    if v is None:
        return None
    return str(v)


def _load_rows_from_stata(file_path: str) -> list[tuple]:
    """Load Stata .dta into (None, headers, rows) tuples."""
    import pyreadstat
    df, _meta = pyreadstat.read_dta(file_path)
    headers = list(df.columns)
    rows = [[_safe_str(v) for v in row] for row in df.values.tolist()]
    return [(None, headers, rows)]


def _load_rows_from_spss(file_path: str) -> list[tuple]:
    """Load SPSS .sav into (None, headers, rows) tuples."""
    import pyreadstat
    df, _meta = pyreadstat.read_sav(file_path)
    headers = list(df.columns)
    rows = [[_safe_str(v) for v in row] for row in df.values.tolist()]
    return [(None, headers, rows)]


def _load_rows_from_rdata(file_path: str) -> list[tuple]:
    """Load R data files into (object_name, headers, rows) tuples."""
    import pyreadr
    result = pyreadr.read_r(file_path)
    tables = []
    for name, df in result.items():
        headers = list(df.columns)
        rows = [[_safe_str(v) for v in row] for row in df.values.tolist()]
        tables.append((name, headers, rows))
    return tables


def _load_tabular_file(file_path: str, ext: str) -> list[tuple]:
    """Dispatch to the correct loader based on extension.

    Returns list of (sheet_or_name, headers, rows) tuples.
    """
    if ext in (".csv", ".tab", ".tsv"):
        return _load_rows_from_csv(file_path, ext)
    if ext in (".xlsx", ".xls"):
        return _load_rows_from_excel(file_path, ext)
    if ext == ".dta":
        return _load_rows_from_stata(file_path)
    if ext == ".sav":
        return _load_rows_from_spss(file_path)
    if ext in (".rds", ".rda"):
        return _load_rows_from_rdata(file_path)
    return []


def _sanitize_column_name(name: str) -> str:
    """Sanitize a column name into a valid SQL identifier."""
    safe = re.sub(r"[^a-zA-Z0-9_]", "_", name).strip("_")
    if not safe or not safe[0].isalpha():
        safe = "col_" + safe
    return safe


def ingest_to_sql(files: list[tuple], documents_dir: str, cfg: dict) -> dict:
    """Ingest tabular files into SQLite and generate schema registry.

    Args:
        files: List of (file_path, dataset_name) tuples.
        documents_dir: Root documents directory.
        cfg: App config dict.

    Returns:
        Schema registry dict (also saved to sql_schemas.json).
    """
    sql_db_dir = cfg.get("paths", {}).get("sql_db", "sql_db")
    if not os.path.isabs(sql_db_dir):
        project_root = Path(__file__).resolve().parent.parent
        sql_db_dir = os.path.join(str(project_root), sql_db_dir)

    os.makedirs(sql_db_dir, exist_ok=True)
    db_path = os.path.join(sql_db_dir, "knowledge_base.db")
    schema_path = os.path.join(sql_db_dir, "sql_schemas.json")

    # Filter to tabular files only
    tabular_files = [(fp, ds) for fp, ds in files if fp.suffix.lower() in SQL_EXTENSIONS]

    if not tabular_files:
        if os.path.exists(schema_path):
            os.remove(schema_path)
        return {}

    # Clear existing DB
    if os.path.exists(db_path):
        os.remove(db_path)

    conn = sqlite3.connect(db_path)
    schema_registry = {}

    try:
        schema_registry = _ingest_tables(conn, tabular_files, documents_dir)
    finally:
        conn.close()

    # Enrich with LLM-generated column descriptions
    if schema_registry:
        try:
            _enrich_schema_with_descriptions(
                schema_registry, tabular_files, documents_dir, cfg,
            )
        except Exception as e:
            print(f"  Column description generation failed (non-fatal): {e}")

    with open(schema_path, "w", encoding="utf-8") as f:
        json.dump(schema_registry, f, indent=2, default=str)

    return schema_registry


def _ingest_tables(
    conn: sqlite3.Connection,
    tabular_files: list[tuple],
    documents_dir: str,
) -> dict:
    """Load tabular files into SQLite tables. Returns the schema registry."""
    schema_registry = {}

    for file_path, dataset_name in tabular_files:
        ext = file_path.suffix.lower()
        try:
            rel_path = file_path.relative_to(documents_dir)
        except ValueError:
            print(f"  SQL ingest: {file_path.name} not under {documents_dir}, skipping")
            continue
        source_file = str(rel_path)

        try:
            tables = _load_tabular_file(str(file_path), ext)
        except Exception as e:
            print(f"  SQL ingest error for {file_path.name}: {e}")
            continue

        for sheet_or_name, headers, rows in tables:
            table_name = _sanitize_table_name(dataset_name, file_path.stem, ext, sheet_or_name)

            # Infer column types, collect samples and stats
            safe_headers = [_sanitize_column_name(h) for h in headers]
            # Deduplicate: ensure all column names are unique after sanitization.
            # Checks every name against all previously assigned names,
            # handling both same-name collisions and suffix collisions
            # (e.g., 'Score_A', 'Score_A', 'Score_A_2' won't produce duplicates).
            used: set[str] = set()
            for i, name in enumerate(safe_headers):
                if name.lower() in used:
                    suffix = 2
                    while f"{name}_{suffix}".lower() in used:
                        suffix += 1
                    safe_headers[i] = f"{name}_{suffix}"
                used.add(safe_headers[i].lower())
            col_types = []
            col_samples = []
            col_stats = []
            for col_idx in range(len(headers)):
                col_values = [row[col_idx] if col_idx < len(row) else None for row in rows]
                ctype = _infer_column_type(col_values)
                col_types.append(ctype)
                col_samples.append(_get_sample_values(col_values))
                col_stats.append(_get_column_stats(col_values, ctype))

            # Create table
            col_defs = ", ".join(f'"{h}" {t}' for h, t in zip(safe_headers, col_types))
            conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
            conn.execute(f'CREATE TABLE "{table_name}" ({col_defs})')

            # Insert rows
            placeholders = ", ".join(["?"] * len(safe_headers))
            insert_sql = f'INSERT INTO "{table_name}" VALUES ({placeholders})'

            for row in rows:
                values = []
                for col_idx, col_type in enumerate(col_types):
                    raw = row[col_idx] if col_idx < len(row) else None
                    if raw is None or str(raw).strip() == "" or str(raw).lower() == "nan":
                        values.append(None)
                    elif col_type == "INTEGER":
                        try:
                            values.append(int(float(str(raw).strip())))
                        except (ValueError, TypeError):
                            values.append(None)
                    elif col_type == "REAL":
                        try:
                            values.append(float(str(raw).strip()))
                        except (ValueError, TypeError):
                            values.append(None)
                    else:
                        values.append(str(raw).strip())
                conn.execute(insert_sql, values)

            conn.commit()

            row_count = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]

            schema_registry[table_name] = {
                "source_file": source_file,
                "columns": [
                    {
                        "name": safe_headers[i],
                        "original_name": headers[i],
                        "type": col_types[i],
                        "sample": col_samples[i],
                        "stats": col_stats[i],
                    }
                    for i in range(len(headers))
                ],
                "row_count": row_count,
            }
            print(f"  SQL: {table_name} ({row_count} rows, {len(headers)} columns)")

    return schema_registry
